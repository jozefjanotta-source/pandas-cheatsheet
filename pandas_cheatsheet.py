# ==========================================
# PANDAS CHEAT SHEET (WORK VERSION)
# ==========================================
# ==========================================
# 🐼 PANDAS CHEAT SHEET (WORK VERSION)
# ==========================================


# ==========================================
# IMPORTS
# ==========================================

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import glob
import os


# ==========================================
# FILE HANDLING
# ==========================================

file_name = os.path.basename(file)


# ==========================================
# LOAD MULTIPLE EXCEL FILES
# ==========================================

files = glob.glob("data/*.xlsx")
df = pd.concat([pd.read_excel(f) for f in files])


# ==========================================
# FILTERING & SELECTION
# ==========================================

df[df["col"] > 0]
df[(df["a"] > 0) & (df["b"] < 10)]

df = df[df["Plant"] == "DE01"]
df = df[(df["Stock"] > 0) & (df["Blocked"] == 0)]

df = df[["Material", "Plant", "Stock", "Value"]]   # reduce columns


# ==========================================
# CALCULATED COLUMNS
# ==========================================

df["Stock_Value"] = df["Stock"] * df["Price"]

df["Category"] = np.where(
    df["Stock"] > 1000,
    "High",
    "Low"
)

df.loc[df["a"] > 5, "b"] = -1


# ==========================================
# CLEANING
# ==========================================

df = df.fillna(0)
df = df.dropna()
df = df.drop_duplicates()


# ==========================================
# DATE & TIME
# ==========================================

df["date"] = pd.to_datetime(df["date"])

df["date"].dt.year
df["date"].dt.month

df["date"].min()
df["date"].max()


# ==========================================
# MERGE / JOIN
# ==========================================

df = pd.merge(df1, df2, on="key", how="left")

df = pd.merge(
    stock,
    material_data,
    on="material",
    how="left"
)


# ==========================================
# CONCAT
# ==========================================

df = pd.concat([df1, df2])
df = pd.concat([df1, df2], keys=["A", "B"])


# ==========================================
# SORTING
# ==========================================

df = df.sort_values("date")


# ==========================================
# GROUPBY (CORE SKILL)
# ==========================================

df.groupby("col")["val"].sum()

df.groupby("col").agg({
    "val": ["mean", "max"]
})


# ==========================================
# PIVOT / RESHAPE
# ==========================================

df.pivot(index="a", columns="b", values="c")

df.pivot_table(
    values="value",
    index="location",
    columns="parameter",
    aggfunc="mean"
)

df.melt(id_vars="id")


# ==========================================
# CHAINING (POWER MOVE)
# ==========================================

df = (
    df[df["Stock"] > 0]
    .merge(material_master, on="Material", how="left")
    .groupby("Plant")["Stock"].sum()
    .reset_index()
)


# ==========================================
# PERFORMANCE
# ==========================================

df.query("col > 5")
df.eval("a + b")

# AVOID:
# df.iterrows()
# df.apply(axis=1)


# ==========================================
# I/O
# ==========================================

pd.read_csv("file.csv")
pd.read_parquet("file.parquet")

df.to_csv("file.csv", index=False)
df.to_parquet("out.parquet")


# ==========================================
# STRING OPS
# ==========================================

df["col"].str.lower()
df["col"].str.contains("x")
df["col"].str.replace("a", "b")


# ==========================================
# TIME SERIES
# ==========================================

df.set_index("date", inplace=True)

df.resample("M").mean()
df.rolling(30).mean()


# ==========================================
# RANK / TOP N
# ==========================================

df.sort_values("col")
df.nlargest(10, "col")

df["rank"] = df["col"].rank()


# ==========================================
# PLOTTING
# ==========================================

df = pd.read_csv("file.csv", index_col=0, parse_dates=True)

df.plot()
plt.show()

df["column"].plot()
plt.show()

df.plot.scatter(x="col1", y="col2", alpha=0.5)
plt.show()

df.plot(kind="bar")
plt.show()

df.plot.box()
plt.show()

df.plot.area(figsize=(12, 4), subplots=True)
plt.show()


# ==========================================
# EXCEL FORMATTING (openpyxl)
# ==========================================

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule

wb = load_workbook("file.xlsx")
ws = wb.active

ws["A1"].font = Font(bold=True, color="FF0000")
ws["A1"].fill = PatternFill(start_color="FFFF00", fill_type="solid")
ws["A1"].alignment = Alignment(horizontal="center")

thin = Side(style="thin")
ws["A1"].border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws["B2"].number_format = "#,##0.00"
ws.freeze_panes = "I2"

rule = FormulaRule(
    formula=['$E2="AVAILABLE_STOCK"'],
    font=Font(bold=True)
)

ws.conditional_formatting.add("A2:Z100", rule)

wb.save("output.xlsx")